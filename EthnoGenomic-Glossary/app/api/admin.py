import io
import json
from typing import List

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook

from app.db.db import get_db
from app.dependencies import require_editor, require_admin
from app.models.term import Term
from app.schemas.term import TermCreate, TermOut, TermUpdate
from app.services import term_service

router = APIRouter(prefix="/admin", tags=["admin"])


@router.get("/terms", response_model=List[TermOut])
def list_terms(db: Session = Depends(get_db), _: str = Depends(require_editor)):
    return db.query(Term).all()


@router.post("/term", response_model=TermOut, status_code=status.HTTP_201_CREATED)
def create_term_admin(term: TermCreate, db: Session = Depends(get_db), _: str = Depends(require_editor)):
    return term_service.create_term(db, term)


@router.put("/terms/{term_id}", response_model=TermOut)
def update_term(term_id: int, payload: TermUpdate, db: Session = Depends(get_db), _: str = Depends(require_editor)):
    term = term_service.get_term(db, term_id)
    if not term:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Term not found")
    for field, value in payload.dict(exclude_unset=True).items():
        setattr(term, field, value)
    db.add(term)
    db.commit()
    db.refresh(term)
    return term


@router.delete("/terms/{term_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_term(term_id: int, db: Session = Depends(get_db), _: str = Depends(require_admin)):
    term = term_service.get_term(db, term_id)
    if not term:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Term not found")
    db.delete(term)
    db.commit()
    return None


@router.post("/import/json", response_model=List[TermOut])
async def import_terms_json(file: UploadFile = File(...), db: Session = Depends(get_db), _: str = Depends(require_editor)):
    content = await file.read()
    try:
        data = json.loads(content)
        if not isinstance(data, list):
            raise ValueError("Expected list of terms")
    except Exception as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Invalid JSON: {exc}")

    created = []
    for item in data:
        term = TermCreate(**item)
        created.append(term_service.create_term(db, term))
    return created


@router.post("/import/xls", response_model=List[TermOut])
async def import_terms_xls(file: UploadFile = File(...), db: Session = Depends(get_db), _: str = Depends(require_editor)):
    content = await file.read()
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    required = ["term_ru", "term_en", "definition", "context", "definition_en", "context_en", "abbreviation", "active"]
    missing = [h for h in required if h not in headers]
    if missing:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Missing columns: {', '.join(missing)}")

    created = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        term_data = {k: row_dict.get(k) for k in required}
        if term_data.get("active") is None:
            term_data["active"] = True
        term = TermCreate(**term_data)
        created.append(term_service.create_term(db, term))
    return created


@router.get("/templates/json")
def download_json_template(_: str = Depends(require_editor)):
    example = [
        {
            "term_ru": "Пример RU",
            "term_en": "Example EN",
            "definition": "Определение на русском",
            "context": "Контекст на русском",
            "definition_en": "Definition in English",
            "context_en": "Context in English",
            "abbreviation": "ABBR",
            "active": True,
        }
    ]
    buf = io.BytesIO(json.dumps(example, ensure_ascii=False, indent=2).encode("utf-8"))
    return StreamingResponse(buf, media_type="application/json", headers={"Content-Disposition": 'attachment; filename="terms_template.json"'})


@router.get("/templates/xls")
def download_xls_template(_: str = Depends(require_editor)):
    wb = Workbook()
    ws = wb.active
    headers = ["term_ru", "term_en", "definition", "context", "definition_en", "context_en", "abbreviation", "active"]
    ws.append(headers)
    ws.append(["Пример RU", "Example EN", "Определение", "Контекст", "Definition", "Context", "ABBR", True])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="terms_template.xlsx"'},
    )
